from rest_framework import viewsets, status
from django.shortcuts import get_object_or_404
from rest_framework.decorators import api_view, permission_classes, action, throttle_classes
from rest_framework.permissions import IsAuthenticated, AllowAny, BasePermission, SAFE_METHODS
from rest_framework.response import Response
from rest_framework.throttling import AnonRateThrottle
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import get_user_model
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from django.http import HttpResponse
from django.db import transaction
from decimal import Decimal
from datetime import date, timedelta
from openpyxl import Workbook, load_workbook
import io

from .models import (Ticket, Tecnico, Opinion, Proveedor, SolicitudServicio,
                     Notificacion, ComentarioTicket, TokenVerificacion, ConfiguracionEmpresa,
                     CodigoRecuperacion, UnidadFlotilla)
from .emails import enviar_html, email_nueva_solicitud, email_confirmacion_cliente, email_codigo_verificacion, email_nuevo_comentario, email_tecnico_asignado
from .serializers import (
    UsuarioSerializer, RegistroSerializer, TicketSerializer,
    TecnicoSerializer, OpinionSerializer, ProveedorSerializer,
    SolicitudServicioSerializer, NotificacionSerializer, ComentarioTicketSerializer,
    ConfiguracionEmpresaSerializer, UnidadFlotillaSerializer
)

User = get_user_model()


def generar_pdf_ticket(ticket):
    """Genera un comprobante de servicio en PDF (reportlab) para un Ticket."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                             topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                             leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('Titulo', parent=styles['Heading1'], textColor=colors.HexColor('#0d0d0d'))
    subtitulo_style = ParagraphStyle('Subtitulo', parent=styles['Normal'], textColor=colors.HexColor('#6b7280'), fontSize=10)

    elementos = []
    elementos.append(Paragraph('Taller<font color="#e85d04">diesel</font> — Comprobante de servicio', titulo_style))
    elementos.append(Paragraph(f'Ticket {ticket.ticket_id}', subtitulo_style))
    elementos.append(Spacer(1, 0.6 * cm))

    datos = [
        ['Empresa', ticket.empresa or '—'],
        ['Fecha', str(ticket.fecha)],
        ['Estatus', ticket.get_estatus_display()],
        ['Tipo de unidad', ticket.tipo_unidad or '—'],
        ['Unidad', ticket.unidad or '—'],
        ['Lugar', ticket.lugar or '—'],
        ['Operador', ticket.operador or '—'],
        ['Técnico asignado', ticket.tecnico.nombre if ticket.tecnico else 'Sin asignar'],
        ['Coordinador', ticket.coordinador.nombre_completo if ticket.coordinador else 'Sin asignar'],
    ]
    tabla_datos = Table(datos, colWidths=[5 * cm, 10 * cm])
    tabla_datos.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#111827')),
        ('FONTSIZE', (0, 0), (-1, -1), 9.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor('#e5e7eb')),
    ]))
    elementos.append(tabla_datos)
    elementos.append(Spacer(1, 0.6 * cm))

    elementos.append(Paragraph('Descripción del servicio', styles['Heading3']))
    elementos.append(Paragraph(ticket.reparacion or 'Sin descripción registrada.', styles['Normal']))
    elementos.append(Spacer(1, 0.6 * cm))

    elementos.append(Paragraph('Total del servicio', styles['Heading3']))
    fmt = lambda n: f'${float(n or 0):,.2f}'
    totales = [
        ['Subtotal', fmt(ticket.total)],
        ['IVA (16%)', fmt(ticket.iva)],
        ['Total a pagar', fmt(ticket.total_f)],
        ['Factura', ticket.factura or 'Sin registrar'],
        ['Estatus de factura', dict(Ticket.EST_FACTURA_CHOICES).get(ticket.estatus_factura, ticket.estatus_factura or '—')],
    ]
    tabla_totales = Table(totales, colWidths=[5 * cm, 10 * cm])
    tabla_totales.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (0, 2), (1, 2), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor('#e5e7eb')),
    ]))
    elementos.append(tabla_totales)
    elementos.append(Spacer(1, 1 * cm))
    elementos.append(Paragraph('Gracias por confiar en TallerDiesel — Auxilio Carretero.', subtitulo_style))

    doc.build(elementos)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{ticket.ticket_id}.pdf"'
    return response


# ── Límites de solicitudes (fuerza bruta) ────────────────────────────────────

class LoginRateThrottle(AnonRateThrottle):
    scope = 'login'


class CodigoRateThrottle(AnonRateThrottle):
    scope = 'auth_codigo'


# Ventana de validez de los códigos de verificación / recuperación
MINUTOS_VALIDEZ_CODIGO = 15
MAX_INTENTOS_CODIGO = 5


# ── Permisos personalizados ───────────────────────────────────────────────────

class EsAdminOPermisoEdicion(BasePermission):
    """Acceso de escritura para Administrador siempre, y para Coordinador
    solo si el Administrador le otorgó el permiso 'puede_editar_sistema'.
    Lectura abierta a cualquier usuario autenticado."""

    def has_permission(self, request, view):
        if not (request.user and request.user.is_authenticated):
            return False
        if request.method in SAFE_METHODS:
            return True
        return request.user.rol == 'admin' or (
            request.user.rol == 'coordinador' and request.user.puede_editar_sistema)


class SoloAdmin(BasePermission):
    """Acceso de escritura exclusivo del Administrador (no delegable)."""

    def has_permission(self, request, view):
        if not (request.user and request.user.is_authenticated):
            return False
        if request.method in SAFE_METHODS:
            return True
        return request.user.rol == 'admin'


class LecturaPublicaEdicionConPermiso(BasePermission):
    """Catálogos (técnicos, proveedores): lectura pública, escritura solo
    para Administrador o Coordinador con permiso de edición otorgado."""

    def has_permission(self, request, view):
        if request.method in SAFE_METHODS:
            return True
        if not (request.user and request.user.is_authenticated):
            return False
        return request.user.rol == 'admin' or (
            request.user.rol == 'coordinador' and request.user.puede_editar_sistema)


class PuedeEditarTicket(BasePermission):
    """El ticket es visible para Cliente, Coordinador y Administrador, pero
    solo Coordinador y Administrador pueden modificarlo o eliminarlo libremente.
    El Cliente nunca puede crear tickets manualmente (se generan automáticamente
    al solicitar un servicio), pero sí puede editar unidad/tipo de unidad/
    descripción de su propio ticket y comentarlo, mientras no esté finalizado."""

    def has_permission(self, request, view):
        if not (request.user and request.user.is_authenticated):
            return False
        if request.method in SAFE_METHODS:
            return True
        if view.action == 'create':
            return request.user.rol in ('admin', 'coordinador')
        return True  # objeto se valida en has_object_permission

    def has_object_permission(self, request, view, obj):
        if request.method in SAFE_METHODS:
            return obj.cliente_id == request.user.id or request.user.rol in ('admin', 'coordinador')
        if request.user.rol in ('admin', 'coordinador'):
            return True
        if obj.cliente_id != request.user.id:
            return False
        # Cliente: solo su propio ticket
        if view.action in ('update', 'partial_update', 'agregar_comentario', 'asignar_tecnico'):
            return obj.estatus != 'terminado'
        if view.action == 'calificar':
            return obj.estatus == 'terminado'
        return False


# ── Auth ──────────────────────────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([LoginRateThrottle])
def login_view(request):
    username = request.data.get('username')
    password = request.data.get('password')
    try:
        user = User.objects.get(username=username)
        if not user.is_active:
            return Response({'error': 'Usuario inactivo.'}, status=400)
        if not user.check_password(password):
            raise Exception()
        if not user.email_verificado:
            return Response({'error': 'Debes verificar tu correo electrónico antes de iniciar sesión.', 'no_verificado': True}, status=403)
        refresh = RefreshToken.for_user(user)
        return Response({
            'access': str(refresh.access_token),
            'refresh': str(refresh),
            'user': UsuarioSerializer(user).data,
        })
    except Exception:
        return Response({'error': 'Credenciales incorrectas.'}, status=400)


@api_view(['POST'])
def logout_view(request):
    """Invalida el refresh token en el servidor (no solo en el navegador),
    para que un token robado o copiado deje de servir de inmediato."""
    try:
        refresh_token = request.data.get('refresh')
        if refresh_token:
            token = RefreshToken(refresh_token)
            token.blacklist()
    except Exception:
        pass  # si el token ya era inválido o expiró, no pasa nada
    return Response({'mensaje': 'Sesión cerrada.'})


@api_view(['POST'])
@permission_classes([AllowAny])
def registro_view(request):
    serializer = RegistroSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        # Crear token de verificación y generar código de 6 dígitos
        token_obj = TokenVerificacion.objects.create(usuario=user)
        codigo = token_obj.generar_codigo()
        token_obj.save(update_fields=['codigo', 'intentos'])
        # Siempre imprimir en consola para desarrollo/debug
        print(f'\n🔑 CÓDIGO DE VERIFICACIÓN para {user.email}: {codigo}\n')
        correo_enviado = False
        try:
            correo_enviado = enviar_html(
                asunto='TallerDiesel — Código de verificación',
                destinatarios=[user.email],
                html_body=email_codigo_verificacion(user.nombre or user.username, codigo, MINUTOS_VALIDEZ_CODIGO),
                text_body=(f'Hola {user.nombre or user.username},\n\n'
                           f'Tu código de verificación es:\n\n{codigo}\n\n'
                           f'Este código es válido por {MINUTOS_VALIDEZ_CODIGO} minutos.\n\n'
                           f'Ingresa este código en la página de verificación para activar tu cuenta.\n\n'
                           f'Si no creaste esta cuenta, ignora este correo.'),
                fail_silently=False,
            )
        except Exception as e:
            print(f'⚠️  Error enviando correo: {e}')
        return Response({
            'mensaje': f'Registro exitoso. {"Revisa tu correo " + user.email + " y escribe el código de verificación." if correo_enviado else "No se pudo enviar el correo. Usa el código que aparece en la terminal del servidor."}',
            'email': user.email,
            'codigo': codigo if not correo_enviado else None,
        }, status=201)
    return Response(serializer.errors, status=400)


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([CodigoRateThrottle])
def verificar_email(request):
    email = (request.data.get('email') or '').strip().lower()
    codigo = (request.data.get('codigo') or '').strip()
    if not email or not codigo:
        return Response({'error': 'Correo y código son requeridos.'}, status=400)
    try:
        user = User.objects.get(email__iexact=email)
        token_obj = TokenVerificacion.objects.get(usuario=user)

        if timezone.now() - token_obj.creado_en > timedelta(minutes=MINUTOS_VALIDEZ_CODIGO):
            token_obj.delete()
            return Response({'error': 'El código expiró. Solicita uno nuevo.', 'expirado': True}, status=400)

        if token_obj.intentos >= MAX_INTENTOS_CODIGO:
            token_obj.delete()
            return Response({'error': 'Demasiados intentos fallidos. Solicita un nuevo código.', 'expirado': True}, status=400)

        if token_obj.codigo != codigo:
            token_obj.intentos += 1
            token_obj.save(update_fields=['intentos'])
            restantes = MAX_INTENTOS_CODIGO - token_obj.intentos
            return Response({'error': f'Código incorrecto. Te quedan {restantes} intento(s).'}, status=400)

        user.email_verificado = True
        user.save(update_fields=['email_verificado'])
        token_obj.delete()
        return Response({'mensaje': 'Correo verificado correctamente. Ya puedes iniciar sesión.'})
    except (User.DoesNotExist, TokenVerificacion.DoesNotExist):
        return Response({'error': 'Código inválido o ya utilizado.'}, status=400)


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([CodigoRateThrottle])
def reenviar_verificacion(request):
    email = request.data.get('email')
    try:
        user = User.objects.get(email__iexact=email, email_verificado=False)
        token_obj, _ = TokenVerificacion.objects.get_or_create(usuario=user)
        codigo = token_obj.generar_codigo()
        token_obj.save(update_fields=['codigo', 'intentos'])
        print(f'\n🔑 CÓDIGO DE VERIFICACIÓN (reenvío) para {user.email}: {codigo}\n')
        correo_enviado = False
        try:
            correo_enviado = enviar_html(
                asunto='TallerDiesel — Código de verificación',
                destinatarios=[user.email],
                html_body=email_codigo_verificacion(user.nombre or user.username, codigo, MINUTOS_VALIDEZ_CODIGO),
                text_body=f'Hola {user.nombre or user.username},\n\nTu nuevo código de verificación es:\n\n{codigo}\n\nVálido por {MINUTOS_VALIDEZ_CODIGO} minutos.',
                fail_silently=False,
            )
        except Exception as e:
            print(f'⚠️  Error enviando correo (reenvío): {e}')
        return Response({
            'mensaje': f'Código reenviado. {"Revisa tu correo " + user.email + "." if correo_enviado else "No se pudo enviar el correo. Usa el código que aparece en la terminal del servidor."}',
            'email': user.email,
            'codigo': codigo if not correo_enviado else None,
        })
    except User.DoesNotExist:
        return Response({'error': 'No se encontró usuario con ese correo o ya está verificado.'}, status=400)


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([CodigoRateThrottle])
def solicitar_recuperacion(request):
    """Olvidé mi contraseña: envía un código de 6 dígitos al correo del usuario."""
    email = (request.data.get('email') or '').strip().lower()
    if not email:
        return Response({'error': 'El correo es requerido.'}, status=400)
    try:
        user = User.objects.get(email__iexact=email)
    except User.DoesNotExist:
        # No revelamos si el correo existe o no, por seguridad
        return Response({'mensaje': 'Si el correo existe, te enviamos un código de recuperación.'})

    codigo_obj, _ = CodigoRecuperacion.objects.get_or_create(usuario=user)
    codigo = codigo_obj.generar_codigo()
    codigo_obj.save(update_fields=['codigo', 'intentos'])
    print(f'\n🔑 CÓDIGO DE RECUPERACIÓN para {user.email}: {codigo}\n')
    try:
        send_mail(
            subject='TallerDiesel — Recupera tu contraseña',
            message=(f'Hola {user.nombre or user.username},\n\n'
                     f'Tu código para restablecer tu contraseña es:\n\n{codigo}\n\n'
                     f'Este código es válido por {MINUTOS_VALIDEZ_CODIGO} minutos.\n\n'
                     f'Si no solicitaste este cambio, ignora este correo.'),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=True,
        )
    except Exception as e:
        print(f'⚠️  Error enviando correo de recuperación: {e}')
    return Response({'mensaje': 'Si el correo existe, te enviamos un código de recuperación.'})


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([CodigoRateThrottle])
def confirmar_recuperacion(request):
    """Verifica el código y establece la nueva contraseña."""
    email    = (request.data.get('email') or '').strip().lower()
    codigo   = (request.data.get('codigo') or '').strip()
    password = request.data.get('password') or ''
    if not email or not codigo or not password:
        return Response({'error': 'Correo, código y nueva contraseña son requeridos.'}, status=400)
    if len(password) < 8:
        return Response({'error': 'La contraseña debe tener al menos 8 caracteres.'}, status=400)
    try:
        user = User.objects.get(email__iexact=email)
        codigo_obj = CodigoRecuperacion.objects.get(usuario=user)

        if timezone.now() - codigo_obj.creado_en > timedelta(minutes=MINUTOS_VALIDEZ_CODIGO):
            codigo_obj.delete()
            return Response({'error': 'El código expiró. Solicita uno nuevo.', 'expirado': True}, status=400)

        if codigo_obj.intentos >= MAX_INTENTOS_CODIGO:
            codigo_obj.delete()
            return Response({'error': 'Demasiados intentos fallidos. Solicita un nuevo código.', 'expirado': True}, status=400)

        if codigo_obj.codigo != codigo:
            codigo_obj.intentos += 1
            codigo_obj.save(update_fields=['intentos'])
            restantes = MAX_INTENTOS_CODIGO - codigo_obj.intentos
            return Response({'error': f'Código incorrecto. Te quedan {restantes} intento(s).'}, status=400)

        user.set_password(password)
        user.save(update_fields=['password'])
        codigo_obj.delete()
        return Response({'mensaje': 'Contraseña actualizada correctamente. Ya puedes iniciar sesión.'})
    except (User.DoesNotExist, CodigoRecuperacion.DoesNotExist):
        return Response({'error': 'Código inválido o ya utilizado.'}, status=400)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def me_view(request):
    return Response(UsuarioSerializer(request.user).data)


# ── Usuarios ──────────────────────────────────────────────────────────────────

class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().order_by('-fecha_registro')
    serializer_class = UsuarioSerializer
    permission_classes = [EsAdminOPermisoEdicion]

    def get_queryset(self):
        qs = super().get_queryset()
        if self.request.user.rol == 'cliente':
            qs = qs.filter(id=self.request.user.id)
        rol = self.request.query_params.get('rol')
        if rol:
            qs = qs.filter(rol=rol)
        return qs

    def _bloquear_permiso_si_no_admin(self, request):
        # Solo el Administrador puede otorgar/revocar el permiso de edición
        # de un Coordinador. Un Coordinador con ese permiso no puede
        # otorgárselo a otros ni a sí mismo.
        if 'puede_editar_sistema' in request.data and request.user.rol != 'admin':
            request.data.pop('puede_editar_sistema')

    def create(self, request, *args, **kwargs):
        if request.user.rol != 'admin' and request.data.get('rol') == 'admin':
            return Response({'error': 'Solo el Administrador puede crear cuentas de Administrador.'}, status=403)
        self._bloquear_permiso_si_no_admin(request)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        objetivo = self.get_object()
        if request.user.rol != 'admin' and objetivo.rol == 'admin':
            return Response({'error': 'No puedes modificar una cuenta de Administrador.'}, status=403)
        if request.user.rol != 'admin' and request.data.get('rol') == 'admin':
            return Response({'error': 'Solo el Administrador puede asignar el rol de Administrador.'}, status=403)
        self._bloquear_permiso_si_no_admin(request)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        objetivo = self.get_object()
        if request.user.rol != 'admin' and objetivo.rol == 'admin':
            return Response({'error': 'No puedes modificar una cuenta de Administrador.'}, status=403)
        if request.user.rol != 'admin' and request.data.get('rol') == 'admin':
            return Response({'error': 'Solo el Administrador puede asignar el rol de Administrador.'}, status=403)
        self._bloquear_permiso_si_no_admin(request)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        objetivo = self.get_object()
        if request.user.rol != 'admin' and objetivo.rol == 'admin':
            return Response({'error': 'No puedes eliminar una cuenta de Administrador.'}, status=403)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def coordinadores(self, request):
        # Admin ve coordinadores + admins; coordinador solo se ve a sí mismo
        if request.user.rol == 'admin':
            qs = User.objects.filter(rol__in=['coordinador', 'admin'], is_active=True)
        else:
            qs = User.objects.filter(rol='coordinador', is_active=True)
        return Response(UsuarioSerializer(qs, many=True).data)


# ── Tickets ───────────────────────────────────────────────────────────────────

class TicketViewSet(viewsets.ModelViewSet):
    queryset = Ticket.objects.all()
    serializer_class = TicketSerializer
    permission_classes = [PuedeEditarTicket]

    def get_queryset(self):
        from django.db.models import Q
        qs = super().get_queryset()
        # Por default no se muestran los tickets en la papelera, salvo que se pida explícitamente
        # con ?eliminados=1 (usado por la acción "papelera" de abajo).
        if self.request.query_params.get('eliminados') == '1':
            qs = qs.filter(eliminado=True)
        else:
            qs = qs.filter(eliminado=False)
        user = self.request.user
        if user.rol == 'cliente':
            qs = qs.filter(cliente=user)
        elif user.rol == 'coordinador':
            # Ve sus propios tickets + tickets sin coordinador asignado (pendientes de atender)
            qs = qs.filter(Q(coordinador=user) | Q(coordinador__isnull=True))
        estatus = self.request.query_params.get('estatus')
        if estatus:
            qs = qs.filter(estatus=estatus)
        return qs

    def perform_destroy(self, instance):
        """No se borra el registro de verdad — se manda a la papelera, para
        poder consultarlo después y para que su folio (TDxxx) quede libre y
        el siguiente ticket nuevo lo reutilice."""
        from django.utils import timezone
        instance.eliminado = True
        instance.eliminado_en = timezone.now()
        instance.eliminado_por = self.request.user
        instance.save()

    @action(detail=False, methods=['get'])
    def papelera(self, request):
        """Lista de tickets eliminados (solo Admin y Coordinador)."""
        if request.user.rol not in ('admin', 'coordinador'):
            return Response({'error': 'No tienes permiso para ver la papelera.'}, status=403)
        qs = Ticket.objects.filter(eliminado=True).order_by('-eliminado_en')
        return Response(TicketSerializer(qs, many=True).data)

    @action(detail=True, methods=['post'])
    def restaurar(self, request, pk=None):
        """Regresa un ticket eliminado a la lista activa (solo Admin y Coordinador)."""
        if request.user.rol not in ('admin', 'coordinador'):
            return Response({'error': 'No tienes permiso para restaurar tickets.'}, status=403)
        ticket = get_object_or_404(Ticket, pk=pk, eliminado=True)
        # Si mientras estuvo en la papelera ya se reutilizó su folio en un ticket nuevo,
        # se le agrega un sufijo para no romper el folio único entre los activos.
        if Ticket.objects.filter(ticket_id=ticket.ticket_id, eliminado=False).exclude(pk=ticket.pk).exists():
            ticket.ticket_id = f'{ticket.ticket_id}-R'
        ticket.eliminado = False
        ticket.eliminado_en = None
        ticket.eliminado_por = None
        ticket.save()
        return Response(TicketSerializer(ticket).data)

    def update(self, request, *args, **kwargs):
        ticket = self.get_object()

        # El cliente solo puede modificar unidad, tipo de unidad y descripción
        if request.user.rol == 'cliente':
            campos_permitidos = {'unidad', 'tipo_unidad', 'reparacion'}
            data = request.data
            if hasattr(data, 'dict'):
                data = data.dict()
            campos_no_permitidos = set(data.keys()) - campos_permitidos
            if campos_no_permitidos:
                return Response(
                    {'error': f'Solo puedes editar unidad, tipo de unidad y descripción. Campos no permitidos: {", ".join(campos_no_permitidos)}.'},
                    status=403)
            kwargs['partial'] = True

        estatus_anterior = ticket.estatus
        tecnico_anterior_id = ticket.tecnico_id
        response = super().update(request, *args, **kwargs)
        ticket.refresh_from_db()

        # Registrar cambio de estatus en bitácora
        if ticket.estatus != estatus_anterior:
            label_anterior = dict(Ticket.ESTATUS_CHOICES).get(estatus_anterior, estatus_anterior)
            ComentarioTicket.objects.create(
                ticket=ticket,
                autor=request.user,
                autor_nombre=request.user.nombre_completo,
                texto=f'Estatus cambiado de "{label_anterior}" a "{ticket.get_estatus_display()}"',
                es_cambio_estatus=True,
                estatus_anterior=estatus_anterior,
                estatus_nuevo=ticket.estatus,
            )
            # Notificar al cliente que su ticket cambió de estatus
            if ticket.cliente_id:
                Notificacion.objects.create(
                    destinatario=ticket.cliente,
                    titulo=f'Tu servicio {ticket.ticket_id} fue actualizado',
                    mensaje=f'El estatus cambió de "{label_anterior}" a "{ticket.get_estatus_display()}".',
                    tipo='cambio_estatus',
                    referencia_id=ticket.id,
                )
            # Notificar al coordinador asignado (si no es él quien hizo el cambio)
            if ticket.coordinador_id and ticket.coordinador_id != request.user.id:
                Notificacion.objects.create(
                    destinatario=ticket.coordinador,
                    titulo=f'Ticket {ticket.ticket_id} actualizado',
                    mensaje=f'Estatus cambiado a "{ticket.get_estatus_display()}" por {request.user.nombre_completo}.',
                    tipo='cambio_estatus',
                    referencia_id=ticket.id,
                )

        # Registrar asignación de técnico en bitácora
        if ticket.tecnico_id and ticket.tecnico_id != tecnico_anterior_id:
            ComentarioTicket.objects.create(
                ticket=ticket,
                autor=request.user,
                autor_nombre=request.user.nombre_completo,
                texto=f'Técnico asignado: {ticket.tecnico.nombre}',
            )
        return response

    @action(detail=True, methods=['post'])
    def asignar_tecnico(self, request, pk=None):
        """El Coordinador o el Administrador asignan un técnico al ticket.
        El propio Cliente también puede fijarlo, pero solo en su propio ticket,
        recién creado y sin técnico todavía (elección hecha en el mapa).
        Al asignarlo, si el ticket estaba 'Pendiente' pasa automáticamente
        a 'Atendido'."""
        ticket = self.get_object()
        es_staff = request.user.rol in ('admin', 'coordinador')
        # El cliente puede elegir o CAMBIAR el técnico de su propio ticket, siempre que
        # sea un servicio directo con técnico (no uno atendido por coordinador) y no
        # esté ya finalizado.
        es_cliente_dueño_tecnico_directo = (
            request.user.rol == 'cliente'
            and ticket.cliente_id == request.user.id
            and ticket.tipo_solicitud == 'tecnico'
            and ticket.estatus != 'terminado'
        )
        if not (es_staff or es_cliente_dueño_tecnico_directo):
            return Response({'error': 'No tienes permiso para asignar técnicos.'}, status=403)
        ticket = self.get_object()
        tecnico_id = request.data.get('tecnico_id') or request.data.get('tecnico')
        if not tecnico_id:
            return Response({'error': 'tecnico_id es requerido.'}, status=400)
        try:
            tecnico = Tecnico.objects.get(pk=tecnico_id)
        except Tecnico.DoesNotExist:
            return Response({'error': 'Técnico no encontrado.'}, status=404)

        tecnico_anterior_id = ticket.tecnico_id
        es_cambio = tecnico_anterior_id is not None and tecnico_anterior_id != tecnico.id
        ticket.tecnico = tecnico
        estatus_anterior = ticket.estatus
        if ticket.estatus == 'pendiente':
            ticket.estatus = 'atendido'
        ticket.save()

        ComentarioTicket.objects.create(
            ticket=ticket, autor=request.user, autor_nombre=request.user.nombre_completo,
            texto=f'Técnico asignado: {tecnico.nombre}',
        )
        if ticket.estatus != estatus_anterior:
            ComentarioTicket.objects.create(
                ticket=ticket, autor=request.user, autor_nombre=request.user.nombre_completo,
                texto=f'Estatus cambiado de "{dict(Ticket.ESTATUS_CHOICES).get(estatus_anterior)}" a "{ticket.get_estatus_display()}"',
                es_cambio_estatus=True, estatus_anterior=estatus_anterior, estatus_nuevo=ticket.estatus,
            )
        # Si la solicitud de origen aún no estaba marcada como atendida, márcala.
        if hasattr(ticket, 'solicitud_origen') and ticket.solicitud_origen and not ticket.solicitud_origen.atendida:
            ticket.solicitud_origen.atendida = True
            ticket.solicitud_origen.save()

        if ticket.cliente_id:
            Notificacion.objects.create(
                destinatario=ticket.cliente,
                titulo='Técnico actualizado' if es_cambio else 'Técnico asignado a tu servicio',
                mensaje=f'{tecnico.nombre} {"ahora" if es_cambio else ""} fue asignado a tu ticket {ticket.ticket_id}.',
                tipo='tecnico_asignado', referencia_id=ticket.id,
            )
            if ticket.cliente.email:
                try:
                    url = f'{settings.FRONTEND_URL}/cliente/dashboard'
                    asunto = f'{"🔧 Cambiamos tu técnico" if es_cambio else "🔧 Ya tienes técnico asignado"} — {ticket.ticket_id}'
                    texto_plano = f'{tecnico.nombre} {"ahora" if es_cambio else ""} atenderá tu servicio (ticket {ticket.ticket_id}).\n\nIngresa al sistema: {url}'
                    enviar_html(asunto, [ticket.cliente.email],
                                email_tecnico_asignado(tecnico.nombre, ticket.ticket_id, es_cambio, url),
                                texto_plano)
                except Exception as e:
                    print(f'⚠️  Error enviando correo de técnico asignado: {e}')

        return Response(TicketSerializer(ticket).data)

    @action(detail=True, methods=['post'])
    def tomar_como_coordinador(self, request, pk=None):
        """Cuando un cliente pide en un comentario que su servicio (originalmente
        directo con técnico) lo atienda un coordinador, este endpoint permite que
        el Coordinador (o Admin, indicando a quién) tome el ticket. A partir de
        ahí el ticket muestra la línea de tiempo por estatus como cualquier
        servicio coordinado."""
        ticket = self.get_object()
        if request.user.rol not in ('admin', 'coordinador'):
            return Response({'error': 'No tienes permiso para tomar este servicio.'}, status=403)
        if ticket.estatus == 'terminado':
            return Response({'error': 'Este ticket ya está finalizado.'}, status=400)

        coordinador_id = request.data.get('coordinador_id')
        if coordinador_id:
            try:
                nuevo_coordinador = User.objects.get(pk=coordinador_id, rol='coordinador')
            except Usuario.DoesNotExist:
                return Response({'error': 'Coordinador no encontrado.'}, status=404)
        elif request.user.rol == 'coordinador':
            nuevo_coordinador = request.user
        else:
            return Response({'error': 'Indica qué coordinador va a tomar este servicio.'}, status=400)

        ticket.coordinador = nuevo_coordinador
        ticket.save()

        ComentarioTicket.objects.create(
            ticket=ticket, autor=request.user, autor_nombre=request.user.nombre_completo,
            texto=f'{nuevo_coordinador.nombre_completo} tomó este servicio como coordinador.',
        )

        if ticket.cliente_id:
            Notificacion.objects.create(
                destinatario=ticket.cliente,
                titulo='Un coordinador tomó tu servicio',
                mensaje=f'{nuevo_coordinador.nombre_completo} ahora coordina tu ticket {ticket.ticket_id}.',
                tipo='coordinador_asignado', referencia_id=ticket.id,
            )
            if ticket.cliente.email:
                try:
                    url = f'{settings.FRONTEND_URL}/cliente/dashboard'
                    enviar_html(
                        asunto=f'🧑‍💼 Un coordinador tomó tu servicio — {ticket.ticket_id}',
                        destinatarios=[ticket.cliente.email],
                        html_body=email_nuevo_comentario(nuevo_coordinador.nombre_completo, 'coordinador', ticket.ticket_id,
                                                          f'{nuevo_coordinador.nombre_completo} tomó tu servicio y ahora le da seguimiento personal.', url),
                        text_body=f'{nuevo_coordinador.nombre_completo} tomó tu servicio {ticket.ticket_id} y ahora le da seguimiento. Ingresa al sistema: {url}',
                    )
                except Exception as e:
                    print(f'⚠️  Error enviando correo de coordinador asignado: {e}')

        return Response(TicketSerializer(ticket).data)

    @action(detail=True, methods=['post'])
    def agregar_comentario(self, request, pk=None):
        ticket = self.get_object()
        texto = request.data.get('texto', '').strip()
        if not texto:
            return Response({'error': 'El comentario no puede estar vacío.'}, status=400)
        comentario = ComentarioTicket.objects.create(
            ticket=ticket,
            autor=request.user,
            autor_nombre=request.user.nombre_completo,
            texto=texto,
        )
        self._notificar_comentario(ticket, request.user, texto)
        return Response(ComentarioTicketSerializer(comentario).data, status=201)

    def _notificar_comentario(self, ticket, autor, texto):
        """Avisa (campana + correo) a la otra parte de la conversación cuando se agrega un comentario."""
        try:
            if autor.rol == 'cliente':
                # Avisar al coordinador asignado, o a los admins si aún no hay coordinador
                destinatarios_usuarios = []
                if ticket.coordinador_id:
                    destinatarios_usuarios = [ticket.coordinador]
                else:
                    destinatarios_usuarios = list(User.objects.filter(rol='admin', is_active=True))
                emails = [u.email for u in destinatarios_usuarios if u.email]
                correo_empresa = settings.EMAIL_EMPRESA
                if correo_empresa not in emails:
                    emails.append(correo_empresa)
                for u in destinatarios_usuarios:
                    Notificacion.objects.create(
                        destinatario=u,
                        titulo='Nuevo mensaje del cliente',
                        mensaje=f'{autor.nombre_completo} comentó en el ticket {ticket.ticket_id}: "{texto[:80]}"',
                        tipo='comentario', referencia_id=ticket.id,
                    )
                if emails:
                    url = f'{settings.FRONTEND_URL}/coordinador/dashboard'
                    enviar_html(
                        asunto=f'💬 Nuevo mensaje del cliente — {ticket.ticket_id}',
                        destinatarios=emails,
                        html_body=email_nuevo_comentario(autor.nombre_completo, 'cliente', ticket.ticket_id, texto, url),
                        text_body=f'{autor.nombre_completo} comentó en el ticket {ticket.ticket_id}:\n\n"{texto}"\n\nIngresa al sistema: {url}',
                    )
            elif autor.rol in ('coordinador', 'admin'):
                if ticket.cliente_id and ticket.cliente.email:
                    Notificacion.objects.create(
                        destinatario=ticket.cliente,
                        titulo='Respuesta a tu solicitud',
                        mensaje=f'{autor.nombre_completo} respondió en tu ticket {ticket.ticket_id}: "{texto[:80]}"',
                        tipo='comentario', referencia_id=ticket.id,
                    )
                    url = f'{settings.FRONTEND_URL}/cliente/dashboard'
                    enviar_html(
                        asunto=f'💬 Respuesta a tu solicitud — {ticket.ticket_id}',
                        destinatarios=[ticket.cliente.email],
                        html_body=email_nuevo_comentario(autor.nombre_completo, 'coordinador', ticket.ticket_id, texto, url),
                        text_body=f'{autor.nombre_completo} respondió en tu ticket {ticket.ticket_id}:\n\n"{texto}"\n\nIngresa al sistema: {url}',
                    )
        except Exception as e:
            print(f'⚠️  Error notificando comentario: {e}')

    @action(detail=True, methods=['post'])
    def calificar(self, request, pk=None):
        """El cliente califica el servicio de un ticket ya finalizado (1-5 estrellas)."""
        ticket = self.get_object()
        if not ticket.tecnico_id:
            return Response({'error': 'Este ticket no tiene técnico asignado, no se puede calificar.'}, status=400)
        if Opinion.objects.filter(ticket=ticket).exists():
            return Response({'error': 'Ya calificaste este servicio.'}, status=400)
        try:
            calificacion = int(request.data.get('calificacion'))
            assert 1 <= calificacion <= 5
        except Exception:
            return Response({'error': 'La calificación debe ser un número entre 1 y 5.'}, status=400)
        opinion = Opinion.objects.create(
            tecnico=ticket.tecnico,
            ticket=ticket,
            nombre_autor=request.user.nombre_completo,
            calificacion=calificacion,
            comentario=(request.data.get('comentario') or '').strip(),
        )
        return Response(OpinionSerializer(opinion).data, status=201)

    @action(detail=True, methods=['get'])
    def pdf(self, request, pk=None):
        """Genera un comprobante de servicio en PDF para el ticket."""
        ticket = self.get_object()
        return generar_pdf_ticket(ticket)

    @action(detail=False, methods=['get'])
    def dashboard_cliente(self, request):
        user = request.user
        tickets = Ticket.objects.filter(cliente=user)
        return Response({
            'en_curso': tickets.filter(estatus__in=['pendiente', 'atendido', 'proceso']).count(),
            'terminados': tickets.filter(estatus='terminado').count(),
            'pagos_pendientes': tickets.filter(estatus_factura='pendiente').count(),
            'total': tickets.count(),
            'recientes': TicketSerializer(tickets, many=True).data,
        })

    @action(detail=False, methods=['get'])
    def dashboard_coordinador(self, request):
        user = request.user
        tickets = Ticket.objects.filter(coordinador=user)
        return Response({
            'en_proceso': tickets.filter(estatus='proceso').count(),
            'pendientes': tickets.filter(estatus__in=['pendiente', 'atendido']).count(),
            'terminados': tickets.filter(estatus='terminado').count(),
            'ganancias_hoy': 0,
            'tickets': TicketSerializer(tickets, many=True).data,
        })

    @action(detail=False, methods=['get'])
    def dashboard_admin(self, request):
        tickets = Ticket.objects.all()
        total_ingresos = sum(float(t.total) for t in tickets)
        total_ganancia = sum(float(t.ganancia) for t in tickets)
        total_iva = sum(float(t.iva) for t in tickets)
        total_comision = sum(float(t.comision) for t in tickets)
        total_costo = sum(float(t.costo) for t in tickets)
        from collections import defaultdict
        empresas = defaultdict(float)
        for t in tickets:
            empresas[t.empresa] += float(t.total)
        top_empresas = sorted(empresas.items(), key=lambda x: -x[1])[:5]

        # Rendimiento por técnico: cuántos tickets, ganancia generada y calificación promedio
        tecnicos_stats = defaultdict(lambda: {'tickets': 0, 'ganancia': 0.0, 'calificaciones': []})
        for t in tickets.filter(tecnico__isnull=False):
            key = t.tecnico.nombre
            tecnicos_stats[key]['tickets'] += 1
            tecnicos_stats[key]['ganancia'] += float(t.ganancia)
        for op in Opinion.objects.filter(tecnico__isnull=False):
            tecnicos_stats[op.tecnico.nombre]['calificaciones'].append(op.calificacion)
        tickets_por_tecnico = [{
            'tecnico': nombre,
            'tickets': s['tickets'],
            'ganancia': s['ganancia'],
            'calificacion_promedio': round(sum(s['calificaciones']) / len(s['calificaciones']), 1) if s['calificaciones'] else None,
        } for nombre, s in sorted(tecnicos_stats.items(), key=lambda x: -x[1]['tickets'])]

        return Response({
            'ingresos_totales': total_ingresos,
            'ganancia_total': total_ganancia,
            'iva_acumulado': total_iva,
            'comisiones': total_comision,
            'costo_total': total_costo,
            'servicios_terminados': tickets.filter(estatus='terminado').count(),
            'en_proceso': tickets.filter(estatus='proceso').count(),
            'pendientes': tickets.filter(estatus__in=['pendiente', 'atendido']).count(),
            'facturas_pagadas': tickets.filter(estatus_factura='pagada').count(),
            'facturas_pendientes': tickets.filter(estatus_factura='pendiente').count(),
            'top_empresas': [{'empresa': e, 'total': t} for e, t in top_empresas],
            'tickets_por_tecnico': tickets_por_tecnico,
        })

    @action(detail=True, methods=['post'])
    def permitir_edicion_coordinador(self, request, pk=None):
        """El Administrador activa/desactiva el permiso para que el Coordinador
        pueda editar un ticket ya cerrado (terminado)."""
        if request.user.rol != 'admin':
            return Response({'error': 'Solo el Administrador puede otorgar este permiso.'}, status=403)
        ticket = self.get_object()
        activar = request.data.get('activar', True)
        ticket.puede_editar_coordinador = bool(activar)
        ticket.save(update_fields=['puede_editar_coordinador'])
        return Response({'puede_editar_coordinador': ticket.puede_editar_coordinador,
                         'ticket_id': ticket.ticket_id})

    @action(detail=False, methods=['get'])
    def corte_mensual(self, request):
        tickets = Ticket.objects.all()
        meses = {}
        for t in tickets:
            key = t.fecha.strftime('%B De %Y').capitalize()
            if key not in meses:
                meses[key] = {'mes': key, 'tickets': 0, 'total': 0, 'ganancia': 0, 'iva': 0, 'comision': 0}
            meses[key]['tickets'] += 1
            meses[key]['total'] += float(t.total)
            meses[key]['ganancia'] += float(t.ganancia)
            meses[key]['iva'] += float(t.iva)
            meses[key]['comision'] += float(t.comision)
        return Response(list(meses.values()))

    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """Exporta a un archivo .xlsx la Bitácora, respetando los mismos
        filtros visibles en la tabla (empresa, estatus, tipo, búsqueda)."""
        if request.user.rol not in ('admin', 'coordinador'):
            return Response({'error': 'No tienes permiso para exportar.'}, status=403)
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        qs = self.get_queryset()
        estatus = request.query_params.get('estatus')
        tipo = request.query_params.get('tipo')
        busqueda = (request.query_params.get('busqueda') or '').lower()
        empresas_param = request.query_params.get('empresas')  # separadas por coma

        if estatus:
            qs = qs.filter(estatus=estatus)
        if tipo and tipo != 'todos':
            qs = qs.filter(tipo_solicitud=tipo)
        if empresas_param:
            lista_empresas = [e for e in empresas_param.split(',') if e]
            qs = qs.filter(empresa__in=lista_empresas)
        if busqueda:
            from django.db.models import Q
            qs = qs.filter(Q(ticket_id__icontains=busqueda) | Q(empresa__icontains=busqueda) | Q(unidad__icontains=busqueda))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bitácora'
        encabezados = ['Ticket', 'Empresa', 'Fecha', 'Estatus', 'Tipo Unidad', 'Unidad', 'Lugar',
                       'Técnico', 'Coordinador', 'Reparación', 'Costo', 'Ganancia', 'Total', 'IVA',
                       'Total Final', 'Factura', 'Estatus Factura']
        ws.append(encabezados)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='111111', end_color='111111', fill_type='solid')

        for t in qs:
            ws.append([
                t.ticket_id, t.empresa, str(t.fecha), t.get_estatus_display(),
                t.tipo_unidad, t.unidad, t.lugar,
                t.tecnico.nombre if t.tecnico else '', t.coordinador.nombre_completo if t.coordinador else '',
                t.reparacion, float(t.costo), float(t.ganancia), float(t.total), float(t.iva),
                float(t.total_f), t.factura, t.estatus_factura,
            ])

        for col in ws.columns:
            longitud = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(40, longitud + 2)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="bitacora_tallerdiesel.xlsx"'
        return response


# ── Técnicos ──────────────────────────────────────────────────────────────────

class TecnicoViewSet(viewsets.ModelViewSet):
    queryset = Tecnico.objects.filter(activo=True)
    serializer_class = TecnicoSerializer
    permission_classes = [LecturaPublicaEdicionConPermiso]

    def get_queryset(self):
        qs = super().get_queryset()
        from django.db.models import Q
        categoria = self.request.query_params.get('categoria')
        ciudad = self.request.query_params.get('ciudad')
        if categoria:
            qs = qs.filter(categoria=categoria)
        if ciudad:
            qs = qs.filter(Q(ciudad__icontains=ciudad) | Q(estado__icontains=ciudad))
        return qs

    @action(detail=False, methods=['get'])
    def plantilla_excel(self, request):
        """Descarga una plantilla de Excel con las columnas correctas para importar técnicos."""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Tecnicos'
        headers = ['nombre', 'categoria', 'ciudad', 'estado', 'direccion', 'telefono', 'latitud', 'longitud', 'descripcion', 'disponible']
        ws.append(headers)
        ws.append(['Juan Pérez', 'motor_diesel', 'Monterrey', 'Nuevo León', 'Av. Constitución 100',
                   '8121234567', 25.6866, -100.3161, 'Ejemplo de descripción', 'si'])
        ws2 = wb.create_sheet('Categorías válidas')
        ws2.append(['valor (usar este texto en la columna "categoria")', 'significa'])
        for val, label in Tecnico.CATEGORIA_CHOICES:
            ws2.append([val, label])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 45

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="plantilla_tecnicos.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'])
    def importar_excel(self, request):
        """Importa técnicos desde un archivo Excel.
        modo='reemplazar' borra todos los técnicos actuales antes de importar.
        modo='agregar' (default) solo agrega/actualiza por nombre.
        """
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'error': 'Sube un archivo Excel (.xlsx).'}, status=400)
        modo = request.data.get('modo', 'agregar')

        try:
            wb = load_workbook(archivo, data_only=True)
            ws = wb.worksheets[0]
        except Exception as e:
            return Response({'error': f'No se pudo leer el archivo. Asegúrate de que sea un .xlsx válido. ({e})'}, status=400)

        filas = list(ws.iter_rows(min_row=2, values_only=True))
        categorias_validas = dict(Tecnico.CATEGORIA_CHOICES)
        creados, errores = [], []

        with transaction.atomic():
            if modo == 'reemplazar':
                Tecnico.objects.all().delete()

            for i, fila in enumerate(filas, start=2):
                if not fila or all(c is None or str(c).strip() == '' for c in fila):
                    continue  # fila vacía, se ignora
                try:
                    datos = (list(fila) + [None] * 10)[:10]
                    nombre, categoria, ciudad, estado, direccion, telefono, latitud, longitud, descripcion, disponible = datos
                    if not nombre or not categoria or not ciudad or not estado or not telefono or latitud is None or longitud is None:
                        errores.append(f'Fila {i}: faltan datos obligatorios (nombre, categoria, ciudad, estado, telefono, latitud, longitud).')
                        continue
                    categoria = str(categoria).strip()
                    if categoria not in categorias_validas:
                        errores.append(f'Fila {i}: categoría "{categoria}" no es válida. Usa una de: {", ".join(categorias_validas.keys())}.')
                        continue
                    disp_txt = str(disponible).strip().lower() if disponible is not None else 'si'
                    disponible_bool = disp_txt not in ('no', 'false', '0', 'ocupado')
                    tecnico, _ = Tecnico.objects.update_or_create(
                        nombre=str(nombre).strip(),
                        defaults=dict(
                            categoria=categoria,
                            ciudad=str(ciudad).strip(),
                            estado=str(estado).strip(),
                            direccion=str(direccion).strip() if direccion else '',
                            telefono=str(telefono).strip(),
                            latitud=float(latitud),
                            longitud=float(longitud),
                            descripcion=str(descripcion).strip() if descripcion else '',
                            disponible=disponible_bool,
                            activo=True,
                        )
                    )
                    creados.append(tecnico.nombre)
                except Exception as e:
                    errores.append(f'Fila {i}: {e}')

        return Response({
            'importados': len(creados),
            'nombres': creados,
            'errores': errores,
            'modo': modo,
            'total_tecnicos': Tecnico.objects.filter(activo=True).count(),
        }, status=201 if creados or not errores else 400)

    @action(detail=True, methods=['post'])
    def actualizar_disponibilidad(self, request, pk=None):
        """El técnico o admin puede actualizar su disponibilidad."""
        tecnico = self.get_object()
        disponible = request.data.get('disponible')
        if disponible is None:
            return Response({'error': 'Campo disponible requerido.'}, status=400)
        tecnico.disponible = bool(disponible)
        tecnico.save(update_fields=['disponible'])
        return Response({'disponible': tecnico.disponible, 'nombre': tecnico.nombre})

    @action(detail=True, methods=['post'], permission_classes=[AllowAny])
    def agregar_opinion(self, request, pk=None):
        tecnico = self.get_object()
        serializer = OpinionSerializer(data={**request.data, 'tecnico': tecnico.id})
        if serializer.is_valid():
            serializer.save()
            opiniones = tecnico.opiniones.all()
            if opiniones.exists():
                tecnico.calificacion = sum(o.calificacion for o in opiniones) / opiniones.count()
                tecnico.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)


# ── Proveedores ───────────────────────────────────────────────────────────────

class ProveedorViewSet(viewsets.ModelViewSet):
    queryset = Proveedor.objects.filter(activo=True)
    serializer_class = ProveedorSerializer
    permission_classes = [EsAdminOPermisoEdicion]


# ── Catálogo de unidades (flotilla) ──────────────────────────────────────────

class UnidadFlotillaViewSet(viewsets.ModelViewSet):
    """Catálogo reutilizable de unidades por usuario: evita volver a capturar
    tipo de unidad / número cada vez que se solicita un servicio."""
    serializer_class = UnidadFlotillaSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return UnidadFlotilla.objects.filter(propietario=self.request.user, activo=True)

    def perform_create(self, serializer):
        serializer.save(propietario=self.request.user)

    def perform_destroy(self, instance):
        # Baja lógica en vez de borrar, por si ya está referenciada en tickets pasados
        instance.activo = False
        instance.save(update_fields=['activo'])


# ── Solicitud pública ─────────────────────────────────────────────────────────


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def lista_empresas(request):
    """Devuelve las empresas únicas registradas en tickets."""
    empresas = Ticket.objects.values_list('empresa', flat=True).distinct().order_by('empresa')
    return Response(list(empresas))


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def crear_proveedor(request):
    """Crea un nuevo proveedor rápidamente desde el modal."""
    nombre = request.data.get('nombre', '').strip()
    if not nombre:
        return Response({'error': 'El nombre es requerido.'}, status=400)
    prov, created = Proveedor.objects.get_or_create(nombre=nombre)
    return Response({'id': prov.id, 'nombre': prov.nombre, 'created': created}, status=201)


@api_view(['POST'])
@permission_classes([AllowAny])
def solicitud_servicio(request):
    serializer = SolicitudServicioSerializer(data=request.data)
    if serializer.is_valid():
        cliente_usuario = request.user if (request.user and request.user.is_authenticated) else None
        solicitud = serializer.save(cliente=cliente_usuario)

        # ── Generación automática del ticket en estado "Pendiente" ──────────
        # El cliente nunca crea el ticket manualmente: se genera al solicitar
        # el servicio, junto con su bitácora.
        # Si el solicitante es coordinador, asignarlo automáticamente al ticket
        coordinador_auto = None
        if cliente_usuario and getattr(cliente_usuario, 'rol', None) == 'coordinador':
            coordinador_auto = cliente_usuario

        # Determinar tipo: si lo envía el coordinador o es ruta coordinador → 'coordinador', else 'tecnico'
        tipo_sol = request.data.get('tipo_solicitud', 'coordinador')

        # Si es tipo tecnico, intentar obtener el tecnico por su ID
        tecnico_obj = None
        if tipo_sol == 'tecnico':
            tecnico_id = request.data.get('tecnico_id')
            if tecnico_id:
                try:
                    from .models import Tecnico
                    tecnico_obj = Tecnico.objects.get(pk=tecnico_id)
                except Exception:
                    pass

        ticket = Ticket.objects.create(
            empresa=solicitud.empresa or (cliente_usuario.empresa if cliente_usuario else '') or solicitud.nombre_completo,
            fecha=date.today(),
            estatus='pendiente',
            tipo_unidad=solicitud.tipo_unidad or '',
            unidad=solicitud.unidad or '—',
            lugar=solicitud.lugar,
            operador=solicitud.nombre_completo,
            reparacion=solicitud.problema,
            cliente=cliente_usuario,
            coordinador=coordinador_auto,
            tipo_solicitud=tipo_sol,
            tecnico=tecnico_obj,
        )
        solicitud.ticket = ticket
        solicitud.save(update_fields=['ticket'])
        ComentarioTicket.objects.create(
            ticket=ticket,
            autor=cliente_usuario,
            autor_nombre=solicitud.nombre_completo,
            texto='Ticket generado automáticamente a partir de la solicitud de servicio.',
        )

        # Notificar a coordinadores y admins (campana + correo)
        destinatarios = User.objects.filter(rol__in=['coordinador', 'admin'], is_active=True)
        emails_destino = [d.email for d in destinatarios if d.email]

        # El correo de la empresa siempre debe recibir la solicitud de coordinador
        correo_empresa = settings.EMAIL_EMPRESA
        if correo_empresa not in emails_destino:
            emails_destino.append(correo_empresa)

        for dest in destinatarios:
            Notificacion.objects.create(
                destinatario=dest,
                titulo='Nueva solicitud de servicio',
                mensaje=f'{solicitud.nombre_completo} ({solicitud.empresa or "Sin empresa"}) necesita ayuda: {solicitud.problema[:80]}...',
                tipo='solicitud',
                referencia_id=solicitud.id,
            )

        # Correo a coordinadores y admins
        if emails_destino:
            tipo_display = 'TÉCNICO' if tipo_sol == 'tecnico' else 'COORDINADOR'
            asunto_interno = f'🔧 Nueva solicitud de {tipo_display} — {solicitud.nombre_completo}'
            texto_plano = (
                f'Se recibió una nueva solicitud de tipo "{tipo_display}":\n\n'
                f'Nombre:   {solicitud.nombre_completo}\n'
                f'Teléfono: {solicitud.telefono}\n'
                f'Empresa:  {solicitud.empresa or "No especificada"}\n'
                f'Problema: {solicitud.problema}\n\n'
                f'Ticket generado: {ticket.ticket_id}\n\n'
                f'Ingresa al sistema: {settings.FRONTEND_URL}/coordinador/dashboard'
            )
            html_body = email_nueva_solicitud(
                tipo_display=tipo_display,
                nombre=solicitud.nombre_completo,
                telefono=solicitud.telefono,
                empresa=solicitud.empresa,
                problema=solicitud.problema,
                lugar=solicitud.lugar,
                tipo_unidad=solicitud.tipo_unidad,
                unidad=solicitud.unidad,
                tecnico_nombre=tecnico_obj.nombre if tecnico_obj else '',
                ticket_id=ticket.ticket_id,
                fecha_str=solicitud.fecha.strftime('%d/%m/%Y %H:%M'),
                url_dashboard=f'{settings.FRONTEND_URL}/coordinador/dashboard',
            )
            enviar_html(asunto_interno, emails_destino, html_body, texto_plano)

        # Correo de confirmación al cliente
        if solicitud.nombre_completo and solicitud.telefono:
            # Buscar si el usuario tiene correo registrado
            cliente_email = None
            if request.user and request.user.is_authenticated and request.user.email:
                cliente_email = request.user.email

            if cliente_email:
                asunto_cliente = '✅ TallerDiesel — Recibimos tu solicitud'
                texto_plano = (
                    f'Hola {solicitud.nombre_completo},\n\n'
                    f'Recibimos tu solicitud de auxilio carretero. Un coordinador se pondrá en contacto contigo '
                    f'en menos de 5 minutos.\n\n'
                    f'Empresa:  {solicitud.empresa or "No especificada"}\n'
                    f'Problema: {solicitud.problema}\n\n'
                    f'Gracias por confiar en TallerDiesel — Auxilio Carretero.\n'
                )
                html_body = email_confirmacion_cliente(
                    nombre=solicitud.nombre_completo,
                    empresa=solicitud.empresa,
                    problema=solicitud.problema,
                )
                enviar_html(asunto_cliente, [cliente_email], html_body, texto_plano)

        return Response({
            'mensaje': 'Solicitud recibida. Te contactaremos en menos de 5 minutos.',
            'ticket_id': ticket.ticket_id,
            'id': ticket.id,
        }, status=201)
    return Response(serializer.errors, status=400)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def solicitudes_pendientes(request):
    solicitudes = SolicitudServicio.objects.filter(atendida=False).order_by('-fecha')
    return Response(SolicitudServicioSerializer(solicitudes, many=True).data)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def marcar_solicitud_atendida(request, pk):
    try:
        solicitud = SolicitudServicio.objects.get(pk=pk)
        solicitud.atendida = True
        solicitud.save()
        return Response({'mensaje': 'Marcada como atendida.'})
    except SolicitudServicio.DoesNotExist:
        return Response({'error': 'No encontrada.'}, status=404)


# ── Notificaciones ────────────────────────────────────────────────────────────

SLA_MINUTOS_PENDIENTE = 20  # minutos que puede estar un ticket "Pendiente" antes de alertar


def revisar_alertas_sla():
    """Crea una notificación (una sola vez por ticket) cuando un servicio
    lleva demasiado tiempo sin ser atendido ('Pendiente')."""
    limite = timezone.now() - timedelta(minutes=SLA_MINUTOS_PENDIENTE)
    atrasados = Ticket.objects.filter(
        estatus='pendiente', alerta_sla_enviada=False, fecha_creacion__lt=limite
    )
    for t in atrasados:
        if t.coordinador_id:
            destinatarios = [t.coordinador]
        else:
            destinatarios = list(User.objects.filter(rol='admin', is_active=True))
        for u in destinatarios:
            Notificacion.objects.create(
                destinatario=u,
                titulo=f'⏰ SLA: Ticket {t.ticket_id} sin atender',
                mensaje=f'El servicio {t.ticket_id} de {t.empresa} lleva más de {SLA_MINUTOS_PENDIENTE} minutos "Pendiente" sin asignar técnico.',
                tipo='alerta_sla',
                referencia_id=t.id,
            )
        t.alerta_sla_enviada = True
        t.save(update_fields=['alerta_sla_enviada'])


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mis_notificaciones(request):
    if request.user.rol in ('admin', 'coordinador'):
        revisar_alertas_sla()
    notifs = Notificacion.objects.filter(destinatario=request.user)
    return Response(NotificacionSerializer(notifs, many=True).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def marcar_leida(request, pk):
    try:
        notif = Notificacion.objects.get(pk=pk, destinatario=request.user)
        notif.leida = True
        notif.save()
        return Response({'ok': True})
    except Notificacion.DoesNotExist:
        return Response({'error': 'No encontrada.'}, status=404)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def marcar_todas_leidas(request):
    Notificacion.objects.filter(destinatario=request.user, leida=False).update(leida=True)
    return Response({'ok': True})


# ── Configuración de la empresa ─────────────────────────────────────────────
# El nombre de la empresa y demás datos generales solo pueden ser editados
# por el Administrador (no es delegable a Coordinador).

@api_view(['GET', 'PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def empresa_config(request):
    config = ConfiguracionEmpresa.obtener()
    if request.method == 'GET':
        return Response(ConfiguracionEmpresaSerializer(config).data)

    if request.user.rol != 'admin':
        return Response({'error': 'Solo el Administrador puede editar la información de la empresa.'}, status=403)

    partial = request.method == 'PATCH'
    serializer = ConfiguracionEmpresaSerializer(config, data=request.data, partial=partial)
    if serializer.is_valid():
        serializer.save(actualizado_por=request.user)
        return Response(serializer.data)
    return Response(serializer.errors, status=400)
